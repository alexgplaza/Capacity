import os
import zipfile
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from flask import Flask, render_template, request, send_file
import matplotlib
matplotlib.use("Agg")  # 👈 Solución para macOS
import matplotlib.pyplot as plt


app = Flask(__name__)

UPLOAD_FOLDER = "static/uploads"
OUTPUT_FOLDER = "static/outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def procesar_excel(filepath):
    """Procesa el Excel y genera gráficos por cada 'Nombre de Despliegue'."""
    df = pd.read_excel(filepath)
    df["Fecha"] = pd.to_datetime(df["Fecha"], utc=True)

    output_excel = os.path.join(OUTPUT_FOLDER, "datos_con_graficos.xlsx")
    wb = load_workbook(filepath)
    ws = wb.create_sheet("Gráficos")

    fila = 1  # Para la posición de las imágenes

    for despliegue, datos in df.groupby("DEPLOYMENT"):
        plt.figure(figsize=(10, 5))
        plt.plot(datos["Fecha"], datos["Valor"], marker="o", linestyle="-", label="Valor", color="blue")

        #if "Valor2" in datos.columns:
        #    plt.plot(datos["Fecha"], datos["Valor2"], marker="s", linestyle="--", label="Valor2", color="red")

        plt.xlabel("Fecha")
        plt.ylabel("Valores")
        plt.title(f"DEPLOYMENT: {despliegue}")
        plt.legend()
        plt.xticks(rotation=45)
        plt.grid()
        plt.tight_layout()

        img_path = os.path.join(OUTPUT_FOLDER, f"grafico_{despliegue}.png")
        plt.savefig(img_path, bbox_inches="tight")
        plt.close()

        img = Image(img_path)
        ws.add_image(img, f"A{fila}")
        fila += 20

    wb.save(output_excel)

    # Crear un archivo ZIP con los resultados
    zip_path = os.path.join(OUTPUT_FOLDER, "resultados.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(output_excel, "datos_con_graficos.xlsx")
        for file in os.listdir(OUTPUT_FOLDER):
            if file.endswith(".png"):
                zipf.write(os.path.join(OUTPUT_FOLDER, file), file)

    return zip_path

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if "file" not in request.files:
            return "No file part"

        file = request.files["file"]
        if file.filename == "":
            return "No selected file"

        if file:
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)
            zip_path = procesar_excel(filepath)
            return send_file(zip_path, as_attachment=True)

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
